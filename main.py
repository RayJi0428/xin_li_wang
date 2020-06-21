import openpyxl  # excel
import sys
import re
import codecs
import os

if len(sys.argv) == 1:
    input_file = input('請將日記帳拖曳至視窗內:')
else:
    input_file = sys.argv[1:]
#input_file = ['d:\\git\\py\\xin_li_wang\\日記帳20200621202503.xlsx']

print(input_file)
input("確認檔案無誤按下ENTER開始執行:")

summary = {}
error_summary = {}

# Step1-先處理所有應收帳款
for file_path in input_file:
    print('開始計算應收...' + str(file_path))

    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]  # 取第一張表

    for ri in range(1, ws.max_row):  # 從第1列跑到第N列
        subject = ws['E' + str(ri)].value  # 科目
        if subject == '應收帳款':
            title = ws['F' + str(ri)].value  # 摘要
            pay = ws['G' + str(ri)].value  # 借出
            if pay > 0:  # 儲存借出款項
                summary[title] = pay

# Step2-處理所有歸還帳款
for file_path in input_file:
    print('開始計算歸還...' + str(file_path))

    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]  # 取第一張表

    for ri in range(1, ws.max_row):  # 從第1列跑到第N列
        subject = ws['E' + str(ri)].value  # 科目
        if subject == '應收帳款':
            title = ws['F' + str(ri)].value  # 摘要
            back = ws['H' + str(ri)].value  # 歸還

            if back > 0:  # 處理歸還款項
                pay_title = title.replace('沖-', '')
                if summary.__contains__(pay_title) == True:
                    summary[pay_title] -= back
                    if summary[pay_title] == 0:
                        del summary[pay_title]
                else:
                    error_summary[pay_title] = back
                    print('[對應問題]: {' + pay_title + '} 歸還 {' + str(back) + "}")

log_str = ''
all_pay = 0
error_back = 0
for data in summary.items():
    if data[1] > 0:
        all_pay += data[1]
        log_str += str(data[0]) + "\t" + str(data[1]) + "\n"

log_str += "總借貸 = " + str(all_pay) + "\n"
log_str += "-------------------------------\n"

for data in error_summary.items():
    if data[1] > 0:
        error_back += data[1]
        log_str += str(data[0]) + "\t" + str(data[1]) + "\n"

log_str += "對應錯誤累積 = " + str(error_back)
gg = os.getcwd()
log_file = codecs.open(os.getcwd() + "\\log.txt", 'w', encoding='utf-8')
log_file.truncate()  # 清空
log_file.write(log_str)
log_file.close()


input("done!")
